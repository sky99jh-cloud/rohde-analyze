"""
DMB 송신기 VM602 텍스트 로그 파싱.
tx-1 / tx-2, pa1~pa5 블록별 I_DRV 행·digital 행에서 값 추출.
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any


def _parse_us_date_from_line(line: str) -> datetime | None:
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", line)
    if not m:
        return None
    month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return datetime(year, month, day)


def _parse_current_line(line: str) -> dict[str, float]:
    """*  I_DRV: 5.03 A   I_1A: 5.17 A   ..."""
    line = re.sub(r"^\*\s*", "", line).strip()
    out: dict[str, float] = {}
    for m in re.finditer(r"(I_DRV|I_\d[A-Z]):\s*([-\d.]+)\s*A", line):
        out[m.group(1)] = float(m.group(2))
    return out


def _parse_digital_line(line: str) -> dict[str, float]:
    """*  digital: PWR_A: 192 W   PWR_B: 213 W   ..."""
    line = re.sub(r"^\*\s*", "", line).strip()
    low = line.lower()
    if "digital:" in low:
        idx = low.index("digital:")
        line = line[idx + len("digital:") :].strip()
    out: dict[str, float] = {}
    for m in re.finditer(
        r"(PWR_[A-Z]+|REFL_OUT|VSWR|V_PHASE|V_DC|I_DC):\s*([-\d.]+)",
        line,
    ):
        out[m.group(1)] = float(m.group(2))
    return out


def parse_dmb_log(txt_path: str) -> dict[str, Any]:
    """
    반환:
    {
        "created_on": datetime | None,
        "tx1": { 1: {"currents": dict, "digital": dict}, 2: ..., ... 5 },
        "tx2": { ... },
    }
    """
    with open(txt_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    created_on: datetime | None = None
    for line in lines[:50]:
        if "Calculated Values" in line and created_on is None:
            created_on = _parse_us_date_from_line(line)

    tx1: dict[int, dict[str, dict[str, float]]] = {}
    tx2: dict[int, dict[str, dict[str, float]]] = {}

    i = 0
    while i < len(lines):
        line = lines[i]
        m = re.search(r"Remarks:\s*tx-(\d+)\s+pa(\d+)", line, re.I)
        if m:
            tx = int(m.group(1))
            pa = int(m.group(2))
            cur_line = lines[i + 1] if i + 1 < len(lines) else ""
            dig_line = ""
            for j in range(i + 2, min(i + 6, len(lines))):
                if "digital:" in lines[j].lower():
                    dig_line = lines[j]
                    break
            currents = _parse_current_line(cur_line)
            digital = _parse_digital_line(dig_line)
            block = {"currents": currents, "digital": digital}
            if tx == 1:
                tx1[pa] = block
            elif tx == 2:
                tx2[pa] = block
        i += 1

    return {
        "created_on": created_on,
        "tx1": tx1,
        "tx2": tx2,
    }


def detect_dmb_excel_kind(excel_path: str) -> str:
    """마지막 시트 A1 텍스트로 'txa' | 'txb' | 'unknown'."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[-1]
        a1 = str(ws.cell(1, 1).value or "")
        u = a1.upper()
        if "TX-A" in u or "TX A" in u.replace("-", " "):
            return "txa"
        if "TX-B" in u or "TX B" in u.replace("-", " "):
            return "txb"
        return "unknown"
    finally:
        wb.close()
