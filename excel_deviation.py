"""
이전 시트(최근 1년) 동일 셀 평균 대비 편차가 임계값 이상이면 빨간 글씨로 표시.
"""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from excel_handler import (
    _find_value_col,
    _normalize,
    _resolve_amp_label_key,
    _SPECIAL_LABEL_MAP,
)

_DEVIATION_THRESHOLD_DEFAULT = 0.2
# FWD(F3), REF(I3), AMP Temp, 특이사항(Shoulder 등)은 50% 이상일 때만 알림
_DEVIATION_THRESHOLD_STRICT = 0.5
_FONT_RED = Font(color="FF0000")


def _normalize_sheet_title(title: str) -> str:
    return re.sub(r"\s*\(\d+\)\s*$", "", str(title).strip()).strip()


def parse_sheet_title_date(title: str) -> datetime | None:
    """시트 이름에서 날짜 추출 (YYYY-MM, YYYY-MM-DD, YYYY-MM_n, YYYY.M 등)."""
    t = _normalize_sheet_title(title)
    m = re.match(r"^(\d{4})-(\d{2})_(\d+)$", t)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), 1)
        except ValueError:
            return None
    m = re.match(r"^(\d{4})-(\d{2})(?:-(\d{2}))?$", t)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        d = int(m.group(3)) if m.group(3) else 1
        try:
            return datetime(y, mo, d)
        except ValueError:
            return None
    m = re.match(r"^(\d{4})\.(\d{1,2})$", t)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), 1)
        except ValueError:
            return None
    return None


def _historical_sheets_for_average(
    workbook: Workbook,
    new_sheet: Worksheet,
    ref_date: datetime,
) -> list[Worksheet]:
    """ref_date(측정일) 기준 최근 365일 이내·측정일 이전 시트. 시트명 날짜 없으면 new_sheet 제외 전체."""
    ref_day = ref_date.replace(hour=0, minute=0, second=0, microsecond=0)
    window_start = ref_day - timedelta(days=365)
    dated: list[tuple[datetime, Worksheet]] = []
    undated: list[Worksheet] = []

    for ws in workbook.worksheets:
        if ws is new_sheet:
            continue
        d = parse_sheet_title_date(ws.title)
        if d is not None:
            d0 = d.replace(hour=0, minute=0, second=0, microsecond=0)
            if window_start <= d0 < ref_day:
                dated.append((d0, ws))
        else:
            undated.append(ws)

    if dated:
        dated.sort(key=lambda x: x[0])
        return [w for _, w in dated]
    return undated


def _cell_numeric_value(ws: Worksheet, row: int, col: int) -> float | None:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _average_historical(
    historical: list[Worksheet],
    row: int,
    col: int,
) -> float | None:
    vals: list[float] = []
    for ws in historical:
        n = _cell_numeric_value(ws, row, col)
        if n is not None:
            vals.append(n)
    if not vals:
        return None
    return sum(vals) / len(vals)


def _pct_diff(new_val: float, avg: float) -> float:
    if avg == 0:
        return 0.0 if new_val == 0 else 1.0
    return abs(new_val - avg) / abs(avg)


def apply_deviation_highlight(
    workbook: Workbook,
    new_sheet: Worksheet,
    ref_date: datetime,
    checks: list[tuple[int, int, str, float]],
    log_callback=None,
) -> list[str]:
    """
    checks: (row, col, 라벨, 임계값 비율 0.2 또는 0.5)
    반환: 해당 임계값 이상 이탈 항목 설명 문자열 리스트.
    """
    def log(msg: str) -> None:
        if log_callback:
            log_callback(msg)

    historical = _historical_sheets_for_average(workbook, new_sheet, ref_date)
    if not historical:
        log("  [평균 비교] 참조할 이전 시트가 없어 편차 검사를 건너뜁니다.")
        return []

    log(
        "  [평균 비교] 참조 시트 "
        f"{len(historical)}개 (기본 20% / FWD·REF·AMP Temp·특이사항 50%)"
    )

    alerts: list[str] = []
    for row, col, label, threshold in checks:
        new_v = _cell_numeric_value(new_sheet, row, col)
        if new_v is None:
            continue
        avg = _average_historical(historical, row, col)
        if avg is None:
            continue
        diff = _pct_diff(new_v, avg)
        if diff >= threshold:
            cell = new_sheet.cell(row=row, column=col)
            cell.font = _FONT_RED
            pct = diff * 100
            msg = f"{label}: 현재 {new_v:g} / 1년 평균 {avg:g} (차이 {pct:.1f}%, 기준 {int(threshold*100)}%)"
            alerts.append(msg)
            log(f"  [이탈≥{int(threshold*100)}%] {msg}")

    return alerts


def collect_rohde_deviation_cells(sheet: Worksheet, parsed: dict) -> list[tuple[int, int, str, float]]:
    """갱신된 숫자 셀 (F3, I3, AMP, DTV 특이사항 값) 좌표·라벨·임계값."""
    checks: list[tuple[int, int, str, float]] = []

    if parsed.get("forward_power") is not None:
        checks.append((3, 6, "F3 Forward Power", _DEVIATION_THRESHOLD_STRICT))
    if parsed.get("reflected_power") is not None:
        checks.append((3, 9, "I3 Reflected Power", _DEVIATION_THRESHOLD_STRICT))

    amp_count = int(parsed.get("amp_count", 2))
    amp_count = max(1, min(8, amp_count))

    for row in range(1, (sheet.max_row or 0) + 1):
        b_cell = sheet.cell(row=row, column=2)
        if b_cell.value is None:
            continue
        key = _resolve_amp_label_key(str(b_cell.value))
        if key is None:
            continue
        thr = (
            _DEVIATION_THRESHOLD_STRICT
            if key == "AMP Temp [°C]"
            else _DEVIATION_THRESHOLD_DEFAULT
        )
        for amp_n in range(1, amp_count + 1):
            amp_data = parsed.get(f"amp{amp_n}", {})
            if not isinstance(amp_data, dict):
                continue
            if amp_data.get(key) is None:
                continue
            col = 2 + amp_n
            checks.append((row, col, f"AMP{amp_n} {key}", thr))

    if amp_count < 6:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row or 1):
            label_cell = None
            for cell in row:
                if cell.value is None:
                    continue
                norm = _normalize(str(cell.value))
                if norm in _SPECIAL_LABEL_MAP:
                    label_cell = cell
                    break
            if label_cell is None:
                continue
            norm = _normalize(str(label_cell.value))
            result_key = _SPECIAL_LABEL_MAP[norm]
            if not parsed.get(result_key):
                continue
            val, _unit = parsed[result_key]
            if val is None:
                continue
            row_num = label_cell.row
            val_col = _find_value_col(sheet, row_num, label_cell.column)
            if val_col is None:
                continue
            checks.append(
                (row_num, val_col, f"특이사항 {result_key}", _DEVIATION_THRESHOLD_STRICT)
            )

    return checks


def collect_dmb_deviation_cells(sheet: Worksheet) -> list[tuple[int, int, str, float]]:
    """DMB: PA1~PA5 × 전류·digital 영역 (4~20행, D~H열). AMP Temp 행만 50%, 나머지 20%."""
    checks: list[tuple[int, int, str, float]] = []
    for row in range(4, 21):
        b = sheet.cell(row=row, column=2).value
        if not b:
            continue
        b_lab = str(b).strip()
        thr = _DEVIATION_THRESHOLD_STRICT if b_lab == "AMP Temp" else _DEVIATION_THRESHOLD_DEFAULT
        for col in range(4, 9):
            pa_lab = sheet.cell(row=3, column=col).value
            pa = str(pa_lab).strip() if pa_lab else f"열{col}"
            coord = sheet.cell(row=row, column=col).coordinate
            checks.append((row, col, f"{pa} / {b_lab} ({coord})", thr))
    return checks
