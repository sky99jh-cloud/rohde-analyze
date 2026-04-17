"""
Excel 처리 모듈.
마지막 시트를 복사하고 파싱된 HTML 데이터로 셀 값을 갱신한다.
"""

import copy
import re
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def detect_excel_tx_kind(excel_path: str) -> str:
    """
    마지막 시트 초반에 'AMP 6' 열 헤더 셀이 있으면 UHDTV 양식(uhdtv), 없으면 DTV(dtv).
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[-1]
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=24):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if str(v).strip() == "AMP 6":
                    return "uhdtv"
        return "dtv"
    finally:
        wb.close()


# ── 레이블 정규화 헬퍼 ──────────────────────────────────────────────────────

def _normalize(text: str) -> str:
    """비교용으로 공백, 특수문자, 대소문자를 통일."""
    return re.sub(r"[\s\[\]°()\.,:]+", "", str(text)).lower()


# AMP 레이블 → 파싱 결과 키 매핑 (정규화 후 비교)
_AMP_LABEL_MAP: dict[str, str] = {
    _normalize("AMP Temp [°C]"):      "AMP Temp [°C]",
    _normalize("V Aux in [V]"):       "V Aux in [V]",
    _normalize("V+ Mon [V]"):         "V+ Mon [V]",
    _normalize("I DC [A]"):           "I DC [A]",
    _normalize("I Pre [A]"):          "I Pre [A]",
    _normalize("V5V ACB [V]"):        "V5V ACB [V]",
    _normalize("V 3V5 [V]"):          "V 3V5 [V]",
    _normalize("V 12Mon [V]"):        "V 12Mon [V]",
    _normalize("V Pre Mon [V]"):      "V Pre Mon [V]",
    _normalize("I PRE [A]"):          "I PRE [A]",
    _normalize("I DRV [A]"):          "I DRV [A]",
    _normalize("I 1A [A]"):           "I 1A [A]",
    _normalize("I 2A [A]"):           "I 2A [A]",
    _normalize("I 3A [A]"):           "I 3A [A]",
    _normalize("I 1B [A]"):           "I 1B [A]",
    _normalize("I 2B [A]"):           "I 2B [A]",
    _normalize("I 3B [A]"):           "I 3B [A]",
    _normalize("Power A [V]"):        "Power A [V]",
    _normalize("Power B [V]"):        "Power B [V]",
    _normalize("Power V Ref [V]"):    "Power V Ref [V]",
    _normalize("Power Out [V]"):      "Power Out [V]",
    _normalize("Reflected Out [V]"):  "Reflected Out [V]",
}

# UHDTV 템플릿 B열(단위 생략)·I Pre / I PRE 구분 — 파서 amp 키로 연결
_EXACT_AMP_LABEL_TO_KEY: dict[str, str] = {
    "AMP Temp": "AMP Temp [°C]",
    "V Aux in": "V Aux in [V]",
    "V+ Mon": "V+ Mon [V]",
    "I DC": "I DC [A]",
    "I Pre": "I Pre [A]",
    "I PRE": "I PRE [A]",
    "V5V ACB": "V5V ACB [V]",
    "V 3V5": "V 3V5 [V]",
    "V 12Mon": "V 12Mon [V]",
    "V Pre Mon": "V Pre Mon [V]",
    "I DRV": "I DRV [A]",
    "I Drv": "I DRV [A]",
    "I 1A": "I 1A [A]",
    "I 2A": "I 2A [A]",
    "I 3A": "I 3A [A]",
    "I 1B": "I 1B [A]",
    "I 2B": "I 2B [A]",
    "I 3B": "I 3B [A]",
    "Power A": "Power A [V]",
    "Power B": "Power B [V]",
    "Power V Ref": "Power V Ref [V]",
    "Power Out": "Power Out [V]",
    "Reflected Out": "Reflected Out [V]",
}


def _resolve_amp_label_key(label: str) -> str | None:
    """B열 텍스트 → 파서 amp 딕셔너리 키."""
    s = str(label).strip()
    if s in _EXACT_AMP_LABEL_TO_KEY:
        return _EXACT_AMP_LABEL_TO_KEY[s]
    return _AMP_LABEL_MAP.get(_normalize(s))


# 특이사항 레이블 → 파싱 결과 키 매핑
_SPECIAL_LABEL_MAP: dict[str, str] = {
    _normalize("Shoulder Distance"): "shoulder_distance",
    _normalize("Shoulder Left"):     "shoulder_left",
    _normalize("Shoulder Right"):    "shoulder_right",
    _normalize("Measured Ripple"):   "measured_ripple",
    _normalize("Measured Group Delay"): "measured_group_delay",
}


# ── 시트 복사 ────────────────────────────────────────────────────────────────

def _copy_sheet(workbook: openpyxl.Workbook, source: Worksheet) -> Worksheet:
    """source 시트를 워크북 끝에 복사하여 반환."""
    new_title = _make_unique_title(workbook, source.title)
    target: Worksheet = workbook.copy_worksheet(source)
    target.title = new_title
    # copy_worksheet 는 끝에 추가되지 않는 경우가 있으므로 순서 보정
    idx = workbook.worksheets.index(target)
    workbook.move_sheet(target, offset=len(workbook.worksheets) - 1 - idx)
    return target


def _ensure_single_sheet_selected(workbook: Workbook, active: Worksheet) -> None:
    """
    저장 시 시트 탭이 여러 개 선택(그룹)된 것처럼 열리지 않도록,
    모든 시트의 tabSelected를 끄고 활성 시트만 선택 상태로 둔다.
    """
    sheets = workbook.worksheets
    try:
        idx = sheets.index(active)
    except ValueError:
        idx = max(0, len(sheets) - 1)
    if workbook.views:
        workbook.views[0].activeTab = idx
    for ws in sheets:
        views = getattr(ws, "views", None)
        if not views or not getattr(views, "sheetView", None):
            continue
        for sv in views.sheetView:
            if sv is not None:
                sv.tabSelected = False
    views = getattr(active, "views", None)
    if views and views.sheetView:
        for sv in views.sheetView:
            if sv is not None:
                sv.tabSelected = True


def _make_unique_title(workbook: openpyxl.Workbook, base: str) -> str:
    """이미 존재하는 시트 이름과 겹치지 않는 이름을 생성."""
    existing = {ws.title for ws in workbook.worksheets}
    if base not in existing:
        return base
    n = 1
    while f"{base} ({n})" in existing:
        n += 1
    return f"{base} ({n})"


# ── 날짜 갱신 ────────────────────────────────────────────────────────────────

def _update_date(sheet: Worksheet, created_on: datetime) -> list[str]:
    """
    시트 상단(1~5행)을 스캔하여 연/월/일 숫자를 가진 셀을 찾아 갱신한다.
    인접 셀의 텍스트('년', '월', '일')를 기준으로 판단한다.
    """
    logs: list[str] = []
    max_col = sheet.max_column or 20

    for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is None:
                continue

            # 오른쪽 셀 텍스트로 연/월/일 판단
            right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
            right_text = str(right_cell.value or "").strip()

            if right_text == "년" and isinstance(cell.value, (int, float)):
                cell.value = created_on.year
                logs.append(f"  날짜 갱신: {cell.coordinate} = {created_on.year}년")

            elif right_text == "월" and isinstance(cell.value, (int, float)):
                cell.value = created_on.month
                logs.append(f"  날짜 갱신: {cell.coordinate} = {created_on.month}월")

            elif right_text == "일" and isinstance(cell.value, (int, float)):
                cell.value = created_on.day
                logs.append(f"  날짜 갱신: {cell.coordinate} = {created_on.day}일")

    return logs


def _update_g2_i2_j2_date_row(sheet: Worksheet, created_on: datetime) -> list[str]:
    """
    2행 G2(연도)·I2(MM월)·J2(DD일)에 HTML Created on 일시를 반영한다.
    템플릿은 월을 '03월', 일을 '23일' 형태로 둔다.
    """
    sheet["G2"].value = created_on.year
    sheet["I2"].value = f"{created_on.month:02d}월"
    sheet["J2"].value = f"{created_on.day:02d}일"
    return [
        f"  측정 일자(2행 G2/I2/J2): {created_on.year}년 {created_on.month:02d}월 {created_on.day:02d}일"
    ]


# ── AMP 데이터 갱신 ──────────────────────────────────────────────────────────

def _update_amp_values(sheet: Worksheet, parsed: dict) -> list[str]:
    """
    B열 레이블에 맞춰 C열부터 AMP 1…N 값을 갱신한다.
    DTV: AMP 2개(C·D열), UHDTV: AMP 6개(C~H열) — parsed[\"amp_count\"]에 따름.
    """
    logs: list[str] = []
    amp_count = int(parsed.get("amp_count", 2))
    amp_count = max(1, min(8, amp_count))

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        b_cell = None
        for cell in row:
            if cell.column == 2:  # B열
                b_cell = cell
                break

        if b_cell is None or b_cell.value is None:
            continue

        key = _resolve_amp_label_key(str(b_cell.value))
        if key is None:
            continue

        row_num = b_cell.row
        for amp_n in range(1, amp_count + 1):
            amp_data: dict = parsed.get(f"amp{amp_n}", {})
            if not isinstance(amp_data, dict):
                continue
            val = amp_data.get(key)
            if val is None:
                continue
            col = 2 + amp_n
            sheet.cell(row=row_num, column=col).value = val
            coord = sheet.cell(row=row_num, column=col).coordinate
            logs.append(f"  AMP{amp_n} [{key}] → {coord} = {val}")

    return logs


# ── 특이사항 갱신 ────────────────────────────────────────────────────────────

def _find_value_col(sheet: Worksheet, row_num: int, start_col: int) -> int | None:
    """
    start_col 오른쪽에서 값(숫자 또는 기존 측정값)이 들어있는 첫 번째 열 번호를 반환.
    병합 셀이나 빈 셀을 건너뛰기 위해 최대 4칸까지 탐색한다.
    """
    for col in range(start_col + 1, start_col + 5):
        cell = sheet.cell(row=row_num, column=col)
        if cell.value is not None:
            return col
    return None


def _update_special_values(sheet: Worksheet, parsed: dict) -> list[str]:
    """
    행 전체를 스캔하여 특이사항 레이블을 찾고 값과 단위를 갱신한다.
    레이블이 A/B 어느 열에 있든(병합 셀 포함) 정상 동작한다.
    레이블 오른쪽의 첫 번째 실제 값 셀을 동적으로 탐색한다.
    """
    logs: list[str] = []

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        label_cell = None
        for cell in row:
            if cell.value is None:
                continue
            norm = _normalize(str(cell.value))
            if norm in _SPECIAL_LABEL_MAP:
                label_cell = cell
                break

        if label_cell is None:
            continue

        norm = _normalize(str(label_cell.value))
        result_key = _SPECIAL_LABEL_MAP[norm]

        value_and_unit = parsed.get(result_key)
        if not value_and_unit:
            continue

        val, unit = value_and_unit
        row_num = label_cell.row

        # 레이블 오른쪽에서 첫 번째 값 셀 위치를 동적으로 탐색
        val_col = _find_value_col(sheet, row_num, label_cell.column)
        if val_col is None:
            logs.append(f"  [경고] 특이사항 [{result_key}] 값 셀을 찾을 수 없음 (행 {row_num})")
            continue
        unit_col = val_col + 1

        if val is not None:
            val_cell = sheet.cell(row=row_num, column=val_col)
            val_cell.value = val
            logs.append(f"  특이사항 [{result_key}] 값 → {val_cell.coordinate} = {val}")

        if unit:
            unit_cell = sheet.cell(row=row_num, column=unit_col)
            unit_cell.value = unit
            logs.append(f"  특이사항 [{result_key}] 단위 → {unit_cell.coordinate} = {unit}")

    return logs


# ── F3 / I3 셀 갱신 ─────────────────────────────────────────────────────────

def _update_power_cells(sheet: Worksheet, parsed: dict) -> list[str]:
    """F3(Forward Power), I3(Reflected Power) 셀을 갱신한다."""
    logs: list[str] = []

    fwd = parsed.get("forward_power")
    ref = parsed.get("reflected_power")

    if fwd is not None:
        sheet["F3"] = fwd
        logs.append(f"  Forward Power → F3 = {fwd} W")

    if ref is not None:
        sheet["I3"] = ref
        logs.append(f"  Reflected Power → I3 = {ref} W")

    return logs


# ── 시트 이름 갱신 (날짜 기반) ───────────────────────────────────────────────

def _make_sheet_name(workbook: openpyxl.Workbook, created_on: datetime) -> str:
    """연·월 기반 시트 이름(YYYY_MM). 동일 월에 여러 시트면 _1, _2 … 로 구분."""
    base = created_on.strftime("%Y_%m")
    existing = {ws.title for ws in workbook.worksheets}
    if base not in existing:
        return base
    n = 1
    while f"{base}_{n}" in existing:
        n += 1
    return f"{base}_{n}"


# ── 공개 함수 ────────────────────────────────────────────────────────────────

def update_excel(excel_path: str, parsed: dict, log_callback=None) -> tuple[str, list[str]]:
    """
    Excel 파일의 마지막 시트를 복사한 뒤 parsed 데이터로 셀을 갱신하고 저장한다.

    Parameters
    ----------
    excel_path : 원본 Excel 파일 경로
    parsed     : html_parser.parse_html() 의 반환값
    log_callback : 진행 메시지를 전달할 콜백 함수 (str → None)

    Returns
    -------
    (저장된 파일 경로, 1년 평균 대비 임계값 이탈 항목 설명 리스트)
    """
    from excel_deviation import apply_deviation_highlight, collect_rohde_deviation_cells

    def log(msg: str):
        if log_callback:
            log_callback(msg)

    log(f"Excel 파일 로드 중: {excel_path}")
    workbook = load_workbook(excel_path)

    source_sheet: Worksheet = workbook.worksheets[-1]
    log(f"원본 시트: '{source_sheet.title}'")

    # 시트 복사
    new_sheet = _copy_sheet(workbook, source_sheet)
    log(f"시트 복사 완료: '{new_sheet.title}'")

    # 시트 이름을 날짜 기반으로 변경
    created_on: datetime | None = parsed.get("created_on")
    if created_on:
        new_title = _make_sheet_name(workbook, created_on)
        new_sheet.title = new_title
        log(f"시트 이름 변경: '{new_title}'")

    # 날짜 갱신
    if created_on:
        logs = _update_date(new_sheet, created_on)
        for msg in logs:
            log(msg)
        logs = _update_g2_i2_j2_date_row(new_sheet, created_on)
        for msg in logs:
            log(msg)

    # F3 / I3 Power 셀 갱신
    logs = _update_power_cells(new_sheet, parsed)
    for msg in logs:
        log(msg)

    # AMP 1 … N 데이터 갱신
    logs = _update_amp_values(new_sheet, parsed)
    for msg in logs:
        log(msg)

    # 특이사항 갱신 (DTV만 — UHDTV는 Non Linear / Linear 항목을 기입하지 않음)
    amp_n = int(parsed.get("amp_count", 2))
    if amp_n < 6:
        logs = _update_special_values(new_sheet, parsed)
        for msg in logs:
            log(msg)
    else:
        log("  특이사항(Non Linear / Linear): UHDTV — 생략")

    # 마지막 시트를 활성 시트로 설정
    workbook.active = new_sheet
    log(f"활성 시트 설정: '{new_sheet.title}'")

    # 새 시트 표시 확대/축소 (저장 후 Excel에서 열 때 적용)
    new_sheet.sheet_view.zoomScale = 85
    log("표시 확대/축소: 85%")

    deviation_alerts: list[str] = []
    ref_date = created_on or datetime.now()
    checks = collect_rohde_deviation_cells(new_sheet, parsed)
    if checks:
        log("━━━ 1년 평균 대비 편차 검사 (FWD·AMP Temp·특이 50% / 그 외 20%, REF(I3) 제외) ━━━")
        deviation_alerts = apply_deviation_highlight(
            workbook, new_sheet, ref_date, checks, log_callback=log
        )

    _ensure_single_sheet_selected(workbook, new_sheet)

    # 저장
    workbook.save(excel_path)
    log(f"저장 완료: {excel_path}")

    return excel_path, deviation_alerts
