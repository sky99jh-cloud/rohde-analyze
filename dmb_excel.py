"""
DMB 송신기 Excel 갱신 — 마지막 시트 복제 후 PA1~PA5 열에 로그 값 반영.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_handler import _copy_sheet, _make_sheet_name


def _clear_manual_cells(sheet: Worksheet) -> list[str]:
    """로그에 없는 셀 — 수동 입력용으로 비움 (E2, G2, D21, D22)."""
    for coord in ("E2", "G2", "D21", "D22"):
        sheet[coord].value = None
    return ["  [수동 입력] E2, G2, D21, D22 비움"]


def _update_dmb_date_top(sheet: Worksheet, created_on: datetime) -> list[str]:
    """1행 F·G·H — 년·월·일 문자열."""
    sheet.cell(row=1, column=6).value = f"{created_on.year}년"
    sheet.cell(row=1, column=7).value = f"{created_on.month:02d}월"
    sheet.cell(row=1, column=8).value = f"{created_on.day:02d}일"
    return [
        f"  측정 일자(1행 F/G/H): {created_on.year}년 {created_on.month:02d}월 {created_on.day:02d}일"
    ]


def _write_pa_columns(
    sheet: Worksheet,
    tx_data: dict[int, dict[str, Any]],
    log: Callable[[str], None],
) -> None:
    """PA1~PA5 → 열 D~H(4~8). I_DRV·I_xA/B 행(4~12), DIGITAL(13~20)."""
    for pa in range(1, 6):
        col = 3 + pa
        if pa not in tx_data:
            log(f"  [경고] tx 데이터에 pa{pa} 없음 — 건너뜀")
            continue
        block = tx_data[pa]
        cur: dict[str, float] = block.get("currents") or {}
        dig: dict[str, float] = block.get("digital") or {}

        for row in range(4, 13):
            b = sheet.cell(row=row, column=2).value
            if not b:
                continue
            key = str(b).strip()
            if key in cur:
                sheet.cell(row=row, column=col).value = cur[key]
                log(f"  PA{pa} [{key}] → {sheet.cell(row=row, column=col).coordinate} = {cur[key]}")

        for row in range(13, 21):
            b = sheet.cell(row=row, column=2).value
            if not b:
                continue
            key = str(b).strip()
            if key in dig:
                sheet.cell(row=row, column=col).value = dig[key]
                log(f"  PA{pa} digital [{key}] → {sheet.cell(row=row, column=col).coordinate} = {dig[key]}")


def update_dmb_excel(
    excel_path: str,
    parsed: dict[str, Any],
    tx_num: int,
    log_callback: Callable[[str], None] | None = None,
) -> str:
    """
    tx_num: 1 → DMB TX-A, 2 → DMB TX-B (`parsed['tx1']` / `parsed['tx2']`).
    """
    def log(msg: str) -> None:
        if log_callback:
            log_callback(msg)

    log(f"DMB Excel 로드: {excel_path}")
    wb = load_workbook(excel_path)
    source = wb.worksheets[-1]
    log(f"원본 시트: '{source.title}'")

    new_sheet = _copy_sheet(wb, source)
    log(f"시트 복사: '{new_sheet.title}'")
    for line in _clear_manual_cells(new_sheet):
        log(line)

    created_on: datetime | None = parsed.get("created_on")
    if created_on:
        new_title = _make_sheet_name(wb, created_on)
        new_sheet.title = new_title
        log(f"시트 이름: '{new_title}'")
        for line in _update_dmb_date_top(new_sheet, created_on):
            log(line)

    tx_data = parsed["tx1"] if tx_num == 1 else parsed["tx2"]
    log(f"TX-{tx_num} PA 블록 갱신")
    _write_pa_columns(new_sheet, tx_data, log)

    wb.active = new_sheet
    log(f"활성 시트: '{new_sheet.title}'")

    new_sheet.sheet_view.zoomScale = 85
    log("표시 확대/축소: 85%")

    wb.save(excel_path)
    log(f"저장 완료: {excel_path}")
    return excel_path
