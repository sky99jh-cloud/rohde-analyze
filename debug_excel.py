"""
특이사항 영역의 실제 셀 위치 확인용 디버그 스크립트.
실행 시 파일 선택 다이얼로그가 열립니다.
"""
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

root = tk.Tk()
root.withdraw()
path = filedialog.askopenfilename(
    title="Excel 파일 선택",
    filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")],
)
if not path:
    print("파일을 선택하지 않았습니다.")
    exit()

wb = load_workbook(path)
ws = wb.worksheets[-1]

print(f"시트 이름: {ws.title}")
print(f"최대 행: {ws.max_row}, 최대 열: {ws.max_column}")
print()

print("=== 셀 내용 (1~40행, 값 있는 셀만) ===")
for row in ws.iter_rows(min_row=1, max_row=40):
    for cell in row:
        if cell.value is not None:
            print(f"  [{cell.coordinate}] col={cell.column} => {repr(cell.value)}")

print()
print("=== 병합 셀 범위 ===")
for merged in ws.merged_cells.ranges:
    print(f"  {merged}")
